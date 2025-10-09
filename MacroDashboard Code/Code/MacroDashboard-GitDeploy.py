import pandas as pd
from fredapi import Fred
import os
from openpyxl import load_workbook
from typing import List, Tuple, Union
from pathlib import Path

# Set up FRED API
api = "958ccd9c67808caf9f941367daf6e812"
fred = Fred(api_key=api)

# Input tickers
monthly_tickers = [
    "SAHMREALTIME",
    "RECPROUSM156N",
    "PCEDGC96",
    "PCESC96",
    "RSAFS",
    "DSPIC96",
    "PSAVERT",
    "TOTALSA",
    "REVOLSL",
    "NONREVSL",
    "DGORDER",
    "INDPRO",
    "TCU",
    "HOUST",
    "PERMIT",
    "HSN1F",
    "EXHOSLUSM495S",
    "PAYEMS",
    "ADPMNUSNERSA",
    "UNRATE",
    "U6RATE",
    "CIVPART",
    "EMRATIO",
    "JTSJOR",
    "JTSHIR",
    "JTSTSR",
    "CPIAUCSL",
    "CPILFESL",
    "PPIFIS",
    "PCEPI",
    "PCEPILFE",
    "UMCSENT",
    "T5YIFR",
    "AHETPI",
    "CSUSHPINSA",
    "DTWEXBGS",
    "IQ",
    "IR"
]

quarterly_tickers = [
    "GDPC1",
    "NGDPSAXDCUSQ",
    "GDPNOW",
    "PCECC96",
    "PCNDGC96",
    "PNFIC1",
    "PRFIC1",
    "GCE",
    "ECIWAG",
    "EXPGS",
    "IMPGS",
    "NETEXP"
]

weekly_tickers = [
    "ICSA",
    "CCSA",
    "MORTGAGE30US"
]

daily_tickers = [
    "T10YIE",
    "DTWEXBGS",
    "FEDFUNDS",
    "DGS2",
    "DGS5",
    "DGS10",
    "DBAA"
]

# Fetch Data and Align Index
def Fetch(ticker_list, fred, freq):
    values_dfs = []
    dates_dfs = []

    # Set Frequency
    f = str(freq).strip().upper()
    if f in ("MONTHLY", "M"):
        period_code = "M"
    elif f in ("QUARTERLY", "Q"):
        period_code = "Q"
    elif f in ("WEEKLY", "W"):
        period_code = "W"
    elif f in ("ANNUAL", "A", "Y"):
        period_code = "A"
    elif f in ("DAILY", "D"):
        period_code = "D"
    else:
        raise ValueError(f"Unsupported frequency: {freq}")

    for ticker in ticker_list:
        try:
            s = fred.get_series(ticker)
        except Exception as e:
            print(f"Warning: couldn't fetch {ticker}: {e}")
            continue

        s.index = pd.to_datetime(s.index)
        df = s.reset_index()
        df.columns = ["release_date", ticker]

        if period_code == "D":
            df["Time"] = df["release_date"].dt.floor("D")
        else:
            df["Time"] = df["release_date"].dt.to_period(period_code).dt.to_timestamp()

        grouped = df.groupby("Time").agg({ticker: "last", "release_date": "last"})
        values_dfs.append(grouped[[ticker]])
        dates_dfs.append(grouped[["release_date"]].rename(columns={"release_date": ticker}))

    if not values_dfs:
        return pd.DataFrame(), pd.DataFrame()

    values = pd.concat(values_dfs, axis=1, sort=True)
    dates = pd.concat(dates_dfs, axis=1, sort=True)
    values.index.name = "Time"
    dates.index.name = "Time"

    return values, dates

monthly_values, monthly_dates = Fetch(monthly_tickers, fred, "M")
quarterly_values, quarterly_dates = Fetch(quarterly_tickers, fred, "Q")
weekly_values, weekly_dates = Fetch(weekly_tickers, fred, "W")
daily_values, daily_dates = Fetch(daily_tickers, fred, "D")

# Save the data table to the folder "Raw Data"
base_dir = Path.cwd().parent.parent
print(base_dir)
raw_location = base_dir / "Raw Data"
raw_location.mkdir(parents=True, exist_ok=True)

monthly_values.to_csv(raw_location / "Monthly_Values.csv")
monthly_dates.to_csv(raw_location / "Monthly_Dates.csv")
quarterly_values.to_csv(raw_location / "Quarterly_Values.csv")
quarterly_dates.to_csv(raw_location / "Quarterly_Dates.csv")
weekly_values.to_csv(raw_location / "Weekly_Values.csv")
weekly_dates.to_csv(raw_location / "Weekly_Dates.csv")
daily_values.to_csv(raw_location / "Daily_Values.csv")
daily_dates.to_csv(raw_location / "Daily_Dates.csv")

# GetData from the csv files
def GetData(Ticker: str, Freq: str, n_lags: int = 4):
    if Freq in ("MONTHLY", "M"):
        period_code = "Monthly"
    elif Freq in ("QUARTERLY", "Q"):
        period_code = "Quarterly"
    elif Freq in ("WEEKLY", "W"):
        period_code = "Weekly"
    elif Freq in ("ANNUAL", "A", "Y"):
        period_code = "Annual"
    elif Freq in ("DAILY", "D"):
        period_code = "Daily"
    else:
        raise ValueError(f"Unsupported frequency: {Freq}")

    base = base_dir / "Raw Data"
    dates_path = os.path.join(base, f"{period_code}_Dates.csv")
    values_path = os.path.join(base, f"{period_code}_Values.csv")

    dates = pd.read_csv(dates_path)
    values = pd.read_csv(values_path)

    if Ticker not in values.columns:
        raise KeyError(f"Ticker '{Ticker}' not found in values file.")

    last_idx = values[Ticker].last_valid_index()
    last_pos = values.index.get_loc(last_idx)

    start_pos = max(0, last_pos - n_lags)
    latest_values = values[Ticker].iloc[start_pos: last_pos + 1].tolist()
    latest_date = dates[Ticker].iloc[last_pos]
    return latest_date, latest_values

# Get the template and location for the updated version
folder = base_dir / "MacroDashboard Versions"
input_xlsx = folder / "Economic Dashboard V1-Template.xlsx"
output_xlsx = folder / "Economic Dashboard V1-Template-UPDATE.xlsx"

# Current csv file is in Timestamp format, so change it to datetime
def _to_python_datetime(dt: Union[pd.Timestamp, str, None]):
    """
    Convert pandas.Timestamp -> python datetime for openpyxl.
    If dt is None or unparseable, return it unchanged.
    """
    if isinstance(dt, pd.Timestamp):
        return dt.to_pydatetime()
    return dt

# Load excel template and select the range of rows to update data
wb = load_workbook(input_xlsx)
ws = wb.active
start_row = 3
max_row = ws.max_row

# write data into the rows, one row each time
def write_panel_for_row(
    row: int,
    ticker_col: str,
    freq_col: str,
    date_col: str,
    present_col: str,
    lag_cols: List[str],
    n_lags: int = 4,
) -> None:
    """
    Write one row for a panel (left or right).
    - ticker_col, freq_col, date_col, present_col: column letters (e.g. "B", "E", "C", "F")
    - lag_cols: list of lag column letters in order [Lag1_col, Lag2_col, ...]
    - n_lags: number of lags expected (defaults to 4)
    """
    # read ticker and freq values from the worksheet
    raw_ticker = ws[f"{ticker_col}{row}"].value
    raw_freq = ws[f"{freq_col}{row}"].value

    # If ticker cell is empty -> skip row
    if raw_ticker is None or str(raw_ticker).strip() == "":
        return

    # normalize strings safely
    ticker = str(raw_ticker).strip()
    freq = str(raw_freq).strip() if raw_freq is not None else "M"

    # get data using your GetData function (assumed defined/imported)
    try:
        latest_date, recent_values = GetData(ticker, freq, n_lags=n_lags)
    except Exception as exc:
        # write error to date cell and skip writing values for this row
        ws[f"{date_col}{row}"].value = f"ERR: {exc}"
        return

    # recent_values expected chronological oldest ... most recent
    values = list(recent_values)

    # Ensure fixed length = n_lags + 1 by padding at front (older side) with None
    expected_len = n_lags + 1
    if len(values) < expected_len:
        pad_len = expected_len - len(values)
        values = [None] * pad_len + values

    # Present is last element, lag1 is second-last, etc.
    present_value = values[-1]
    lag_values = []
    for i in range(1, len(lag_cols) + 1):
        # i=1 -> lag1 => values[-2], general: values[-1 - i]
        idx = -1 - i
        # safe access: if index out of range, return None
        try:
            lag_values.append(values[idx])
        except IndexError:
            lag_values.append(None)

    # write date (convert pandas Timestamp -> python datetime for openpyxl)
    ws[f"{date_col}{row}"].value = _to_python_datetime(latest_date)

    # write present and lag values
    ws[f"{present_col}{row}"].value = present_value
    for col_letter, lag_value in zip(lag_cols, lag_values):
        ws[f"{col_letter}{row}"].value = lag_value


# Output section is the left panel, and other three sections are in the right panel.
# Can automate the identification of columns for each panel
# Left panel: B = ticker, E = freq, C = Latest Date, F = Present, G,H,I,J = Lag1..Lag4
left_panel = {
    "ticker_col": "B",
    "freq_col": "E",
    "date_col": "C",
    "present_col": "F",
    "lag_cols": ["G", "H", "I", "J"],
}

# Right panel: M = ticker, P = freq, N = Latest Date, Q = Present, R,S,T,U = Lag1..Lag4
right_panel = {
    "ticker_col": "M",
    "freq_col": "P",
    "date_col": "N",
    "present_col": "Q",
    "lag_cols": ["R", "S", "T", "U"],
}

# iterate and write
for r in range(start_row, max_row + 1):
    write_panel_for_row(r, **left_panel)
    write_panel_for_row(r, **right_panel)

wb.save(output_xlsx)
