import pandas as pd
import os
from openpyxl import load_workbook
from typing import List, Union, Tuple
from pathlib import Path

# Set Location
base_dir = Path(__file__).resolve().parent.parent.parent
print(base_dir)
raw_location = base_dir / "Raw Data"
raw_location.mkdir(parents=True, exist_ok=True)

# GetData from the csv files
def NormalizeFreq(code: str) -> str:
    """Return canonical single-letter code: 'A','Q','M','W','D' or raise."""
    fm = {
        ("MONTHLY", "M"): "M",
        ("QUARTERLY", "Q"): "Q",
        ("WEEKLY", "W"): "W",
        ("ANNUAL", "A", "Y"): "A",
        ("DAILY", "D"): "D",
    }
    code_up = code.upper()
    for keys, val in fm.items():
        if code_up in keys:
            return val
    raise ValueError(f"Unsupported frequency: {code}")

def GetData(ticker: str, freq: str, n_lags: int = 4):
    period_code=NormalizeFreq(freq)
    base = base_dir / "Raw Data"
    dates_path = os.path.join(base, f"{period_code}_Dates.csv")
    values_path = os.path.join(base, f"{period_code}_Values.csv")

    dates = pd.read_csv(dates_path,index_col=0, parse_dates=True)
    values = pd.read_csv(values_path,index_col=0, parse_dates=True)

    if ticker not in values.columns:
        raise KeyError(f"Ticker '{ticker}' not found in values file.")

    last_idx = values[ticker].last_valid_index()
    last_pos = values.index.get_loc(last_idx)
    latest_date = dates[ticker].iloc[last_pos]
    value_col = values[ticker]
    return latest_date, value_col

def GetLevel(ticker: str, freq: str, n_lags: int = 4):
    latest_date, value_col = GetData(ticker, freq, n_lags)
    last_idx = value_col.last_valid_index()
    last_pos = value_col.index.get_loc(last_idx)
    start_pos = max(0, last_pos - n_lags)
    latest_values = value_col.iloc[start_pos: last_pos + 1].tolist()
    return latest_date, latest_values

def GetDelta(ticker: str, freq: str, agg_freq: str, n_lags: int = 4, pct: bool = True) -> Tuple[str, pd.Series]:
    """
    Return (latest_date, delta_series) where delta_series is the same index as value_col
    containing differences (level diffs) or percent changes depending on `pct`.
    - `freq` is the native frequency of the data stored (e.g., "M" or "MONTHLY").
    - `agg_freq` is the aggregation period for the delta (e.g., "A" for year-over-year).
    - If value_col has a DatetimeIndex, the function prefers date-aware shifts (DateOffset).
    - If value_col has a plain integer index, the function uses integer shifts where we can derive them.
    """
    # integer shifts between canonical freq units when treating the series as regular rows
    # meaning: how many rows apart is one agg_freq period when rows are at 'freq' resolution?
    # e.g. if rows are monthly ('M') and agg_freq is 'A' -> shift = 12
    _INT_SHIFT_TABLE = {
        ("M", "A"): 12,
        ("M", "Q"): 3,
        ("Q", "A"): 4,
        ("Q", "M"): None,  # fractional: 1 quarter != integer months when you treat rows as quarters
        ("W", "M"): None,  # ambiguous unless you assume 4 or 4.345
        ("D", "M"): None,
        # same-to-same:
        ("M", "M"): 1,
        ("Q", "Q"): 1,
        ("A", "A"): 1,
        ("W", "W"): 1,
        ("D", "D"): 1,
    }

    # normalize freq inputs
    freq_code = NormalizeFreq(freq)
    agg_code = NormalizeFreq(agg_freq)

    # Get Data
    latest_date, value_col = GetData(ticker, freq_code, n_lags)
    last_idx = value_col.last_valid_index()
    last_pos = value_col.index.get_loc(last_idx)
    start_pos = max(0, last_pos - n_lags)

    # Normal MoM if the freq is M
    if agg_code == freq_code:
        shift_arg = 1
        if pct:
            result = value_col.pct_change(shift_arg,fill_method=None)
        else:
            result = value_col.diff(shift_arg)
        result = result.iloc[start_pos : last_pos + 1].tolist()
        return latest_date, result

    # Default index is datetime index,so this is the default way of calculating delta
    idx = value_col.index
    if isinstance(idx, pd.DatetimeIndex) or pd.api.types.is_datetime64_any_dtype(idx):

        # choose the calendar offset and use .shift(freq=...) to move values for subtraction
        if freq_code == "M" and agg_code == "A":
            offset = pd.DateOffset(months=12)
        elif freq_code == "M" and agg_code == "Q":
            offset = pd.DateOffset(months=3)
        elif freq_code == "Q" and agg_code == "A":
            offset = pd.DateOffset(months=12)
        elif freq_code == "Q" and agg_code == "M":
            raise ValueError("Unsupported conversion: quarter-index to month-based aggregation when using DatetimeIndex.")
        elif freq_code == "W" and agg_code == "A":
            offset = pd.DateOffset(weeks=52)
        elif freq_code == "D" and agg_code == "A":
            offset = pd.DateOffset(years=1)
        else:
            raise ValueError(f"Unsupported date-based conversion: {freq_code} -> {agg_code}")

        # Move the historical values forward so that prev[t] == value_col[t - offset]
        prev = value_col.shift(freq=offset)
        prev = prev.reindex(value_col.index)

        if pct:
            result = (value_col - prev) / prev
        else:
            result = value_col - prev

        result = result.iloc[start_pos : last_pos + 1].tolist()
        return latest_date, result

    # If not datetime index: try integer shift from table _INT_SHIFT_TABLE
    key = (freq_code, agg_code)
    shift = _INT_SHIFT_TABLE.get(key, None)
    if shift is None:
        if freq_code == "Q" and agg_code == "A":
            shift = 4
        elif freq_code == "M" and agg_code == "A":
            shift = 12
        elif freq_code == "M" and agg_code == "Q":
            shift = 3
        else:
            raise ValueError(f"Unsupported conversion for integer-indexed series: {freq_code} -> {agg_code}")

    # safe integer shift
    shift = int(shift)
    if pct:
        result = value_col.pct_change(shift,fill_method=None)
    else:
        result = value_col.diff(shift)

    result = result.iloc[start_pos : last_pos + 1].tolist()

    return latest_date, result

# Get the template and location for the updated version
folder = base_dir / "MacroDashboard Versions"
input_xlsx = folder / "Economic Dashboard V1-Template.xlsx"
output_xlsx = folder / "Economic Dashboard V1-Template-UPDATE.xlsx"

# Current csv file is in Timestamp format, so change it to datetime
def _to_python_datetime(dt: Union[pd.Timestamp, str, None]):
    """
    Convert pandas.Timestamp -> python datetime for openpyxl.
    If dt is None or unparseable, return it unchanged.
    """
    if isinstance(dt, pd.Timestamp):
        return dt.to_pydatetime()
    return dt

# Load excel template and select the range of rows to update data
wb = load_workbook(input_xlsx)
ws = wb.active
start_row = 3
max_row = ws.max_row

# write data into the rows, one row each time

def WritePanel(
    row: int,
    ticker_col: str,
    freq_col: str,
    date_col: str,
    units_col: str,
    present_col: str,
    lag_cols: List[str],
    n_lags: int = 4,
) -> None:
    """
    Write one row for a panel (left or right).
    - ticker_col, freq_col, date_col, present_col: column letters (e.g. "B", "E", "C", "F")
    - lag_cols: list of lag column letters in order [Lag1_col, Lag2_col, ...]
    - n_lags: number of lags expected (defaults to 4)
    """
    # read ticker and freq values from the worksheet
    raw_ticker = ws[f"{ticker_col}{row}"].value
    raw_freq = ws[f"{freq_col}{row}"].value
    raw_units = ws[f"{units_col}{row}"].value

    # If ticker cell is empty -> skip row
    if raw_ticker is None or str(raw_ticker).strip() == "" \
            or str(raw_freq).strip() == "Freq" \
            or len(str(raw_ticker).strip())>20:
        return

    # normalize strings safely
    ticker = str(raw_ticker).strip()
    freq = str(raw_freq).strip() if raw_freq is not None else "M"
    units = str(raw_units).strip().lower()

    # get data using your GetLevel function (assumed defined/imported)
    if('delta' not in units):
        try:
            latest_date, recent_values = GetLevel(ticker, freq, n_lags=n_lags)
        except Exception as exc:
            # write error to date cell and skip writing values for this row
            ws[f"{date_col}{row}"].value = f"ERR: {exc}"
            return
    else:
        pct="%" in units
        agg=units[units.index("/")+1:units.index("/")+2]
        try:
            latest_date, recent_values = GetDelta(ticker, freq, agg, n_lags=n_lags, pct=pct)
        except Exception as exc:
             # write error to date cell and skip writing values for this row
            ws[f"{date_col}{row}"].value = f"ERR: {exc}"
            return

    # recent_values expected chronological oldest ... most recent
    values = list(recent_values)

    # Ensure fixed length = n_lags + 1 by padding at front (older side) with None
    expected_len = n_lags + 1
    if len(values) < expected_len:
        pad_len = expected_len - len(values)
        values = [None] * pad_len + values

    # Present is last element, lag1 is second-last, etc.
    present_value = values[-1]
    lag_values = []
    for i in range(1, len(lag_cols) + 1):
        # i=1 -> lag1 => values[-2], general: values[-1 - i]
        idx = -1 - i
        try:
            lag_values.append(values[idx])
        except IndexError:
            lag_values.append(None)

    # write date (convert pandas Timestamp -> python datetime for openpyxl)
    ws[f"{date_col}{row}"].value = _to_python_datetime(latest_date)

    # write present and lag values
    ws[f"{present_col}{row}"].value = present_value
    for col_letter, lag_value in zip(lag_cols, lag_values):
        ws[f"{col_letter}{row}"].value = lag_value


# Output section is the left panel, and other three sections are in the right panel.
# Can automate the identification of columns for each panel
# Left panel: B = ticker, E = freq, C = Latest Date, F = Present, G,H,I,J = Lag1..Lag4
left_panel = {
    "ticker_col": "B",
    "freq_col": "E",
    "date_col": "C",
    "units_col": "D",
    "present_col": "F",
    "lag_cols": ["G", "H", "I", "J"],
}

# Right panel: M = ticker, P = freq, N = Latest Date, Q = Present, R,S,T,U = Lag1..Lag4
right_panel = {
    "ticker_col": "M",
    "freq_col": "P",
    "date_col": "N",
    "units_col": "O",
    "present_col": "Q",
    "lag_cols": ["R", "S", "T", "U"],
}

# iterate and write
for r in range(start_row, max_row + 1):
    WritePanel(r, **left_panel)
    WritePanel(r, **right_panel)

wb.save(output_xlsx)