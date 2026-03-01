"""
generate_html.py  —  Static HTML generator for the Macro Dashboard.

Reads Economic Dashboard V1-Template-UPDATE.xlsx directly and renders it as
HTML that faithfully matches the Excel layout: column widths, row heights,
gridlines, number formats, yellow highlights, bold headers, etc.

Run via: python generate_html.py
"""

import math
from pathlib import Path
from datetime import datetime, date, timezone

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR      = Path(__file__).resolve().parent
VERSIONS_DIR  = BASE_DIR / "MacroDashboard Versions"
DOCS_DIR      = BASE_DIR / "Website"
XLSX_PATH     = VERSIONS_DIR / "Economic Dashboard V1-Template-UPDATE.xlsx"

# ---------------------------------------------------------------------------
# Conversion constants
# ---------------------------------------------------------------------------
CHAR_TO_PX        = 6.5    # Excel character-width unit → CSS px
PT_TO_PX          = 4 / 3  # Excel row-height points → px
DEFAULT_COL_CHARS = 8.0    # Excel default column width (char units)
DEFAULT_ROW_PT    = 15.0   # Excel default row height (points)

# ---------------------------------------------------------------------------
# Number formatting  (Excel format → Python display string)
# ---------------------------------------------------------------------------

def _fmt_number(val, fmt: str) -> str:
    """Apply an Excel number-format string to a numeric value."""
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return ""

    # Percentage formats
    if fmt == "0.0%":                           # multiply × 100, 1 decimal
        return f"{val * 100:.1f}%"
    if fmt == "0.00%":
        return f"{val * 100:.2f}%"
    if "\\%" in fmt and fmt.startswith("0"):    # value already in percent
        decimals = fmt.count("0") - 1           # e.g. '0\%' → 0 dec
        return f"{round(val, decimals):.{max(0,decimals)}f}%"

    # Thousands-scaled comma integer  (#,##0,)
    if fmt.startswith("#,##0,") and "%" not in fmt:
        return f"{val / 1000:,.0f}"

    # Accounting / comma integer  (#,##0 or * #,##0)
    if "#,##0" in fmt and "%" not in fmt:
        return f"{val:,.0f}"

    # General / fallback
    if fmt in ("General", "@", "") or not fmt:
        if isinstance(val, int):
            return str(val)
        if abs(val) >= 10_000:
            return f"{val:,.0f}"
        if abs(val) >= 1_000:
            return f"{val:,.1f}"
        # show up to 4 significant figures, strip trailing zeros
        s = f"{val:.4g}"
        return s

    return f"{val:.4g}"   # catch-all


def fmt_cell(cell) -> str:
    """Return the display string for a cell, matching Excel's rendered output."""
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    if isinstance(val, (datetime, date)):
        dt = val if isinstance(val, datetime) else datetime(val.year, val.month, val.day)
        # Use cell number format if it looks like a date pattern
        fmt = cell.number_format or ""
        if "yyyy" in fmt.lower() or "mm" in fmt.lower() or "dd" in fmt.lower():
            # Render as YYYY-MM (month precision is fine for dashboard)
            return dt.strftime("%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        return _fmt_number(val, cell.number_format or "General")
    return str(val)

# ---------------------------------------------------------------------------
# Cell style helpers
# ---------------------------------------------------------------------------

def _argb_to_hex(argb: str) -> str | None:
    """Convert an ARGB hex string (8 chars) to a CSS #rrggbb string, or None."""
    if not argb or len(argb) < 6:
        return None
    rgb = argb[-6:].upper()
    if rgb in ("000000", "FFFFFF"):
        return None
    return "#" + rgb


def cell_bg(cell) -> str | None:
    """Return CSS background-color value, or None."""
    try:
        fill = cell.fill
        if fill.fill_type in (None, "none"):
            return None
        fg = fill.fgColor
        if fg and fg.type == "rgb":
            return _argb_to_hex(fg.rgb)
    except Exception:
        pass
    return None


def cell_font_css(cell) -> dict:
    """Return a dict of font-related CSS properties."""
    props = {}
    try:
        f = cell.font
        if f.bold:
            props["font-weight"] = "bold"
        if f.size:
            # Excel stores in points; keep as pt for screen fidelity
            props["font-size"] = f"{f.size:.0f}pt"
        if f.italic:
            props["font-style"] = "italic"
        if f.color and f.color.type == "rgb":
            c = _argb_to_hex(f.color.rgb)
            if c:
                props["color"] = c
    except Exception:
        pass
    return props


def cell_align(cell) -> str:
    """Return CSS text-align, inferring from value type when not explicit."""
    try:
        h = cell.alignment.horizontal
        if h in ("left", "right", "center"):
            return h
        if h == "general":
            pass  # fall through to type-based default
    except Exception:
        pass
    # Infer from value type
    val = cell.value
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return "right"
    return "left"


def cell_border_css(cell) -> dict:
    """Return border CSS properties for explicit (non-default-gridline) borders."""
    props = {}
    try:
        b = cell.border
        def _side(side, key):
            if side and side.border_style:
                w = "2px" if side.border_style == "medium" else "1px"
                try:
                    c = _argb_to_hex(side.color.rgb) if side.color and side.color.type == "rgb" else None
                except Exception:
                    c = None
                color = c or "#888"
                props[key] = f"{w} solid {color}"
        _side(b.top,    "border-top")
        _side(b.bottom, "border-bottom")
        _side(b.left,   "border-left")
        _side(b.right,  "border-right")
    except Exception:
        pass
    return props

# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

def generate_html() -> str:
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    now_utc  = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    max_col  = ws.max_column   # 21  (A … U)
    max_row  = ws.max_row

    # ── Column widths (proportional %) ──────────────────────────────────────
    col_chars = []
    for i in range(1, max_col + 1):
        letter = get_column_letter(i)
        dim    = ws.column_dimensions.get(letter)
        chars  = (dim.width if dim and dim.width else DEFAULT_COL_CHARS)
        col_chars.append(max(1.0, chars))

    total_chars = sum(col_chars)
    col_pct = [f"{c / total_chars * 100:.3f}%" for c in col_chars]

    # ── Row heights (px) ────────────────────────────────────────────────────
    row_px = {}
    for i in range(1, max_row + 1):
        dim = ws.row_dimensions.get(i)
        pt  = (dim.height if dim and dim.height else DEFAULT_ROW_PT) or DEFAULT_ROW_PT
        row_px[i] = max(16, int(pt * PT_TO_PX))

    # ── Merged-cell map ─────────────────────────────────────────────────────
    # (row, col) → (rowspan, colspan)  for the anchor cell
    # (row, col) → None                for covered cells (skip in HTML)
    merged = {}
    for rng in ws.merged_cells.ranges:
        rs = rng.max_row - rng.min_row + 1
        cs = rng.max_col - rng.min_col + 1
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if r == rng.min_row and c == rng.min_col:
                    merged[(r, c)] = (rs, cs)
                else:
                    merged[(r, c)] = None   # covered → skip

    # ── Render rows ─────────────────────────────────────────────────────────
    rows_html = []
    for r in range(1, max_row + 1):
        h_px    = row_px.get(r, int(DEFAULT_ROW_PT * PT_TO_PX))
        cells_html = []

        for c in range(1, max_col + 1):
            m = merged.get((r, c), (1, 1))   # default: normal cell (1×1)
            if m is None:
                continue                       # covered by a merge → skip

            rowspan, colspan = m
            cell = ws.cell(row=r, column=c)

            # ── value string ────────────────────────────────────────────
            val_str = fmt_cell(cell)

            # ── CSS props ───────────────────────────────────────────────
            css = {}
            bg = cell_bg(cell)
            if bg:
                css["background-color"] = bg
            css.update(cell_font_css(cell))
            css["text-align"] = cell_align(cell)
            css.update(cell_border_css(cell))
            css["vertical-align"] = "middle"
            css["padding"] = "2px 5px"
            # Allow text to wrap for long string cells (e.g. footnotes)
            is_long_text = isinstance(cell.value, str) and len(cell.value) > 40
            css["overflow"] = "visible" if is_long_text else "hidden"
            css["white-space"] = "normal" if is_long_text else "nowrap"

            style = "; ".join(f"{k}:{v}" for k, v in css.items())

            # ── build <td> ──────────────────────────────────────────────
            attrs = [f'style="{style}"']
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')

            cells_html.append(f'    <td {" ".join(attrs)}>{val_str}</td>')

        rows_html.append(
            f'  <tr style="height:{h_px}px">\n'
            + "\n".join(cells_html)
            + "\n  </tr>"
        )

    # ── Column group ────────────────────────────────────────────────────────
    colgroup = "\n".join(
        f'  <col style="width:{w}">' for w in col_pct
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Macro Dashboard</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}

body {{
    font-family: Calibri, 'Segoe UI', Arial, sans-serif;
    font-size: 12pt;
    background: #ffffff;
    color: #000000;
    padding: 8px;
}}

table {{
    border-collapse: collapse;
    width: 100%;
    table-layout: fixed;
}}

td {{
    border: 1px solid #d0d0d0;   /* Excel-style gridlines */
    font-size: 12pt;
}}

.footer {{
    margin-top: 8px;
    font-size: 9pt;
    color: #555;
    border-top: 1px solid #ccc;
    padding-top: 6px;
}}

@media print {{
    body {{ padding: 0; }}
    .footer {{ display: none; }}
}}
</style>
</head>
<body>

<table>
<colgroup>
{colgroup}
</colgroup>
<tbody>
{"".join(chr(10) + row for row in rows_html)}
</tbody>
</table>

<div class="footer">Last updated: {now_utc}</div>

</body>
</html>
"""


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    DOCS_DIR.mkdir(exist_ok=True)
    (DOCS_DIR / ".nojekyll").touch()
    out_path = DOCS_DIR / "index.html"
    html = generate_html()
    out_path.write_text(html, encoding="utf-8")
    print(f"Generated {out_path}  ({len(html):,} bytes)")
